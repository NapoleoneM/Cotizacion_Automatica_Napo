"""Genera un .xlsx listo para importar como pestaña 'Horarios' en el
documento espejo de Google Sheets (Archivo -> Importar -> Insertar como
hoja nueva). No se sube al repo ni se ejecuta en producción: es una
herramienta de una sola vez para arrancar la hoja con un ejemplo real.

Rellena EQUIPO con los nombres/turnos reales (mismos de cargar_equipo.py,
solo el personal que SÍ toma chats — soporte/jefa/presencial no necesitan
seguimiento de horario para efectos de cobertura) y ejemplos de cada color
de la leyenda para que la jefa vea cómo se ve cada estado.

Uso:
    pip install openpyxl
    python tools/generar_plantilla_horarios.py
    -> crea Horarios_plantilla.xlsx en la carpeta del proyecto
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# (nombre, turno) — solo Redes/Página web: Soporte, Jefa y Presencial no
# necesitan que se les rastree el horario para efectos de cobertura.
EQUIPO = [
    ("Estefania", 1), ("Ximena", 1), ("Yesica M", 1), ("Susana", 1),
    ("Santiago", 1), ("Yessika", 1),
    ("Gisela", 2), ("Jennifer", 2), ("Angelica", 2), ("Wilmer", 2),
    ("Juan David", 2), ("Kelly", 2),
    ("Yesid", 3), ("Juliana", 3), ("Alexander", 3),
]

TURNOS_LABEL = {
    1: "1 Turno 8:00am a 4:00pm, Sábado 8:00am a 3:00pm",
    2: "2 Turno 11:00am a 7:00pm, Sábado 10:00am a 5:00pm",
    3: "3 Turno 2:00pm a 9:00pm, Sábado 11:00am a 6:00pm",
}

DIAS = ["Lunes 3", "Martes 4", "Miercoles 5", "Jueves 6", "Viernes 7", "Sabado 8", "Domingo 9"]

# Colores de ejemplo para la leyenda (ARGB sin '#', que es lo que pide openpyxl).
COLOR = {
    "compensatorio": "9FC5E8",       # azul claro
    "ausencia": "F4B9C9",            # rosa
    "cambio de horario": "F9CB9C",   # naranja claro
    "santafe": "FFE599",             # amarillo
    "reparte chats": "B6D7A8",       # verde claro
}
BLANCO = "FFFFFF"
GRIS_ENCABEZADO = "D9D9D9"

# Un ejemplo de cada estado, repartido en personas y días distintos (más
# realista que ponerle 5 excepciones a la misma persona en la misma semana).
# (nombre, índice de día 0=Lunes..6=Domingo) -> estado
EJEMPLOS = {
    ("Ximena", 2): "compensatorio",       # miércoles
    ("Yessika", 1): "ausencia",           # martes
    ("Jennifer", 3): "cambio de horario", # jueves
    ("Wilmer", 4): "santafe",             # viernes
    ("Santiago", 5): "reparte chats",     # sábado
}

wb = Workbook()
ws = wb.active
ws.title = "Horarios"

negrita = Font(bold=True)
centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)


def pintar(celda, hexcolor):
    celda.fill = PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")


# --- Título de la semana (opcional, se muestra al pie del panel) ---
ws["A1"] = "Semana del 3 al 9 de Agosto de 2026"
ws["A1"].font = negrita

# --- Fila de encabezado con los días (fila 3) ---
FILA_ENCABEZADO = 3
for i, dia in enumerate(DIAS):
    c = ws.cell(row=FILA_ENCABEZADO, column=2 + i, value=dia)
    c.font = negrita
    c.alignment = centrado
    pintar(c, GRIS_ENCABEZADO)

# --- Bloques de turno + nombres ---
fila = FILA_ENCABEZADO + 1
turno_actual = None
for nombre, turno in EQUIPO:
    if turno != turno_actual:
        turno_actual = turno
        c = ws.cell(row=fila, column=1, value=TURNOS_LABEL[turno])
        c.font = negrita
        fila += 1
    for i in range(len(DIAS)):
        col = 2 + i
        # Domingo libre por defecto, salvo el primer turno (para que el
        # ejemplo muestre también gente trabajando el domingo).
        if i == 6 and turno != 1:
            continue
        c = ws.cell(row=fila, column=col, value=nombre)
        c.alignment = centrado
        estado = EJEMPLOS.get((nombre, i))
        if estado:
            pintar(c, COLOR[estado])
    fila += 1

fila += 1  # fila en blanco antes de la leyenda

# --- Leyenda: texto del estado + color a la derecha ---
ws.cell(row=fila, column=1, value="Leyenda:").font = negrita
fila += 1
for estado, color in COLOR.items():
    ws.cell(row=fila, column=1, value=estado.capitalize() if estado != "santafe" else "Santafe")
    pintar(ws.cell(row=fila, column=2), color)
    fila += 1

# Ancho de columnas para que se lea cómodo
ws.column_dimensions["A"].width = 42
for i in range(len(DIAS)):
    ws.column_dimensions[chr(ord("B") + i)].width = 14

wb.save("Horarios_plantilla.xlsx")
print("Creado: Horarios_plantilla.xlsx")
print(f"{len(EQUIPO)} personas, {len(COLOR)} ejemplos de estado (uno de cada color de la leyenda)")
